

import openpyxl
from openpyxl import load_workbook
from openpyxl import workbook


class Archivo_excel():
	"""Lee el archivo de Recalculo"""

	def __init__(self, documento):		
	

		self.ruta = documento.name
		self.abrir_documento()
		self.obtener_hojas()



	def abrir_documento(self):
		self.documento_open =  load_workbook(self.ruta)
		print(f"Se Abrio el documento: {self.ruta}")
		




	def obtener_hojas(self):
		"""Obtiene los nombres de las hojas
			del archivo excel dado"""

		self.hojas_nombres = {}

		for hoja_nombre, hojas in  zip(self.documento_open.sheetnames,
									   self.documento_open.worksheets):
			self.hojas_nombres[hoja_nombre] = hojas

		self.hojas_lista = [h for h in self.documento_open.worksheets]

		
		print(f"""El documento tiene: {len(self.hojas_nombres)} hojas 
		nombres: {self.hojas_nombres.keys()}""")
		




	def leer_titulos(self,hoja, linea):
		"""Obtiene titulos de columnas del
			archivo que se abre"""

		self.linea_i = linea		
		self.titulos_columnas = {}
		self.obtener_claves_celdas(hoja)

		#LLama al metodo que obtiene los titulos cccn del IQ
		


		if type(self.hoja) is str:

			hoja = self.hojas_nombres[self.hoja]
			titulos = hoja[self.linea_i]

			for titulo in range(len(titulos)):
				if titulos[titulo].value is not None:
					self.titulos_columnas[titulos[titulo].value] = titulos[titulo]



		elif type(self.hoja) is int:

			hoja = self.hojas_lista[self.hoja]
			titulos = hoja[self.linea_i]


			for titulo in range(len(titulos)):
				if titulos[titulo].value is not None:
					self.titulos_columnas[titulos[titulo].value] = titulos[titulo]



	#esta metodo es llamado por el metodo de leer titulos
	def obtener_claves_celdas(self,hoja):

		self.claves_columnas     = {}
		self.columnas_i = {}

		self.hoja = hoja



		if type(self.hoja) is str :
			columna_max = self.hojas_nombres[self.hoja].max_column
			fila_max 	 = self.hojas_nombres[self.hoja].max_row

			for (celda_text,
				 celda_nombres_clv) in self.titulos_columnas.items():

					celda          = str(celda_nombres_clv).split(".")[-1][:-2]
					self.claves_columnas[celda_text.strip(' ')]  = celda.strip('>')
					#fila_sum    = int(self.linea_i) + 1
					self.columnas_i [celda_text.strip(' ')] = self.linea_i

		elif type(self.hoja) == int and type(self.hoja) != float:
			columna_max = self.hojas_lista[self.hoja].max_column
			fila_max 	 = self.hojas_lista[self.hoja].max_row

			for (celda_text,
				 celda_nombres_clv) in self.titulos_columnas.items():

					celda          = str(celda_nombres_clv).split(".")[-1][:-2]
					self.claves_columnas[celda_text.strip(' ')]  = celda.strip('>')
					#fila_sum    = int(self.linea_i) + 1
					self.columnas_i [celda_text.strip(' ')] = self.linea_i

